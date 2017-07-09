#iteration 3 AUfgabe 1
#1.
## MEtrik:
# H. Großbuchstaben
# H. Kleinbuchstaben
# H. Zahlen
# H. Sonderzeichen

#/* k-neaerst Algoithmus:
#1. Termiantionswert 0.5
#2.Zufällig festlegen der lusterzentren
#3.1. Berechnen des Abstandes zum Clusterzentrum aller Datenpunkte!!
#3.2. Die Werte die einem Zentrum am nächsten sind , diesem Zentrum #zuordnen
#3.3 Berechnung des Mittelwertes für jedes Cluster => neues Cluster #zentrum
#*/
from os import path
import matplotlib.pyplot as plt
from wordcloud import WordCloud
import re
import openpyxl
from openpyxl import Workbook
import random
from pylab import *
import psycopg2
#Open File
clusterbook= Workbook()	
sheet1=clusterbook.create_sheet("Cluster1",0)
sheet1 ['A1'] = "Cluster1"
sheet1 ['B1'] = "Cluster2"
sheet1 ['C1'] = "Cluster3"
sheet1 ['D1'] = "Cluster4"
sheet1 ['E1'] = "Cluster5"
sheet1 ['F1'] = "Cluster6"
#sheet1 ['G1'] = "Cluster7"
#sheet1 ['H1'] = "Cluster8"
#sheet1 ['I1'] = "Cluster9"
#sheet1 ['J1'] = "Cluster10"


try:
        conn = psycopg2.connect("dbname='election' user='postgres' host='localhost' password='passwort'") #stellt Verbindung zur Datenbank her
except:
        print ("connection failed") #Fehlermeldung, falls Verbindung nicht gelingt

cur = conn.cursor()
cur.execute("""SELECT hashtag FROM hashtag""")
all_data =cur.fetchall()
#print(all_data[1])
data_array =[]
for i in range (1,len(all_data)):
	temp = str(all_data[i])
	temp= temp.replace("\ ", "")
	temp =temp[:(len(temp)-3)]
	data_array.append(temp[2:])

#sortiert die Einträge nach Clustern/gleichzeitig erstellt es ein Array mit den Wörtern des Clusters
def clusteration(liste,clusternumber):
	cluster_points=[] 
	hashtag_cluster=[]
	for x in liste:	
		if x[1] == clusternumber:
			cluster_points.append(x[0])
	for y in cluster_points:
		hashtag_cluster.append(data_array[y])
	return hashtag_cluster
#Gebe den Clustern_Centren die Eigenschaft die sie besitzen wieviele Gro?-Kleinbuchstaben und Zahlen
def cluster_center_naming(x1,x2,x3,x4):
	cluster_center= "#"
	cluster_center = cluster_center + "A"*x1
	cluster_center = cluster_center + "a"*x2
	cluster_center = cluster_center + "1"*x3
	cluster_center = cluster_center + "#"*x4
	return cluster_center
#special_chars = '@()[]{}"§$%/=,;:-_<>|\`´*+~#*-&.!?ÄÜÖäüöß'
a1=[0]
a2=[0]
a3=[0]
a4=[0]
b1=[0]
b2=[0]
b3=[0]
b4=[0]


#2.
#gebe jedem Hashtag seine Werte Entsprechend Groß-und Kelinschreibung und Zahlen
for n in range (1,len(data_array)):
	a1.append(len(re.findall(r'[A-Z]',data_array[n])))
	a2.append(len(re.findall(r'[a-z]',data_array[n])))
	a3.append(len(re.findall(r'[0-9]',data_array[n])))
	a4.append(len(re.findall(r'[,":*+#-.!?]',data_array[n])))



#K-nearest Algorithmus
#/* 3.*/
ende = False
grenzwert = 0.2
k = 10
for o in range (1,k):
	b1.append(o)
	b2.append(o)
	b3.append(o)
	b4.append(o)
#kmeans-algotithm
while_counter = 0
while (ende== False):
	punkt_Cluster = []
	abstand = 10000
	for i in range (1,len(data_array)-1):	
		for j in range (0,k):
			x1= a1[i]
			x2= a2[i]
			x3= a3[i]
			x4= a4[i]
			y1= b1[j]
			y2= b2[j]
			y3= b3[j]
			y4= b4[j]
			
			#Abstand zwischen Clusterzentrum und Hashtag
			#3.1.
			dca = ((y1-x1)**2+(y2-x2)**2+(y3-x3)**2+(y4-x4)**2)**(0.5)
		#Zuordnen der hashtags zu einem Cluster
		#3.2.
			if (dca<= abstand):
				abstand = dca
				pc_tuple= (i,j)
				
		punkt_Cluster.append(pc_tuple)
		#print(pc_tuple)
		abstand = 10000		
	mean_x1 =0
	mean_x2 =0
	mean_x3 =0
	mean_x4 =0
	counter =0
	#Mittelwertberechnung der Cluster => setzten eines neuen Mittelwertes & Abgleich ob sich die Zentren um mehr als den festgelegten Grenzwert verschieben
	#3.3.
	for l in range (0,k):
		for m in range (0,len(data_array)-1):
			if (punkt_Cluster[m-1] == (m,l)):
				mean_x1 = mean_x1 + a1[m] 
				mean_x2 = mean_x2 + a2[m]
				mean_x3 = mean_x3 + a3[m]
				mean_x4 = mean_x4 + a4[m]
				counter = counter +1
		if (counter==0):
			ende = True
			
		elif (((b1[l] - mean_x1/counter)**2+(b2[l] - mean_x2/counter)**2+(b3[l] - mean_x3/counter)**2+(b4[l] - mean_x4/counter)**2)**(0.5)<grenzwert): 
#+(b4[l] - mean_x4/counter)**2)< grenzwert):
			ende = True	
			b1[l] = mean_x1/counter
			b2[l] = mean_x2/counter
			b3[l] = mean_x3/counter
			b4[l] = mean_x4/counter
		else:
			b1[l] = mean_x1/counter
			b2[l] = mean_x2/counter
			b3[l] = mean_x3/counter
			b4[l] = mean_x4/counter
			ende = False
		#print(str(b1[l])+str(b2[l])+str(b3[l]) +str(b4[l])+str(counter))
		#b4[l] = mean_x4/counter
		mean_x1 =0
		mean_x2 =0
		mean_x3 =0
		mean_x4 =0	
		counter =0
		while_counter = while_counter + 1

#print(while_counter)
#3.
#Visualisieren des Ergebnisses
cluster_centers= []
cluster=[]
for t in range (0,k):
	cluster.append(clusteration(punkt_Cluster,t))
	cluster_centers.append(cluster_center_naming(int(round(b1[t])),int(round(b2[t])),int(round(b3[t])),int(round(b4[t]))))


count=1
anti_count=0
for d in range (-1,k):
	if cluster[d] !=[]:	
	#if len(re.findall(r'[[A-Z],[a-z],[0-9]]',str(cluster[d])))<1:
		#print(len(re.findall(r'[[A-Z],[a-z],[0-9]]',str(cluster[d]))))
		temp_hash = cluster[d];
		for e in range (0,len(temp_hash)):
			sheet1.cell(row = e+2,column=d+2).value= temp_hash[e]		
		text = str(cluster[d])+""
		#print (str(cluster[d]))
		header = str(cluster_centers[d])+""
		if len(header)>14:
			header= header[:14]+"\n"+header[15:28]
			if len(header)>28:
				header= header[:14]+"\n"+header[15:28]+"\n"+header[29:]
		wordcloud=WordCloud().generate(text)
		plt.imshow(wordcloud, interpolation='bilinear')
		#plt.imshow(wordcloud.recolor(color_func=grey_color_func,random_state=3),interpolation="bilinear")
		plt.title(header)
		plt.axis("off")
			
		ap = subplot((ceil(k/3)),3,count)
		if (count>=k-anti_count-1):
			count= k-anti_count-1
			anti_count= anti_count + 1
		else:
			count = count+1	
		plt.subplots_adjust(hspace=1)
	else:
		anti_count=anti_count+1

#plt.delaxes(ap.flatten()[ceil((d+2)/3),((d+2)%3)])	
#plt.show()
plt.savefig("Hashtag_Cluster_Visualisierung_"+str(k)+"Means")
clusterbook.save('Hashtag_Cluster.xlsx')













#Aufgabe2
#Line Chart
#var = dataset.groupby('Standort').Umsatz.sum()
#fig = pyplot.figure()
#ax1 = fig.add_subplot(1,1,1)
#ax1.set_xlabel('Umsatz')
#ax1.set_ylabel('Standort')
#var.plot(kind='line')
#pyplot.show()

# Histogramm
#fig = pyplot.figure()
#ax = fig.add_subplot(1,1,1)
#ax.hist(dataset['Mitarbeiter'], bins=5, color='#9400D3')
#pyplot.title('Mitarbeiter Verteilung')
#pyplot.xlabel('Verteilung')
#pyplot.ylabel('Anzahl Mitarbeiter')
#pyplot.show()
